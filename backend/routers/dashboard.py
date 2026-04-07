from collections import defaultdict
from datetime import date, datetime, time, timedelta, timezone
from io import BytesIO
from math import radians, sin, cos, asin, sqrt
from typing import Dict, List

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session

from database import get_db
from middleware.auth import get_current_user
from models.attendance import AttendanceSession, AttendancePolicy
from models.hotel import Hotel
from models.finance import ShiftReport
from models.maintenance import MaintenanceReport, MaintenanceStatus
from models.room import Room, RoomStatus
from models.user import User, UserRole
from schemas.dashboard import (
    AttendanceRowOut,
    DashboardCountOut,
    DashboardOverviewOut,
    HotelPerformanceOut,
    RoomUptimeOut,
    TechnicianPerformanceOut,
)

router = APIRouter(prefix="/api/dashboard", tags=["Dashboard"])
ATTENDANCE_RADIUS_METERS = 220
ATTENDANCE_STALE_MINUTES = 12
EARLY_CHECKOUT_WARNING_MINUTES = 10
DEFAULT_CHECKIN_START = time(hour=7, minute=0)
DEFAULT_CHECKIN_END = time(hour=10, minute=0)
DEFAULT_SHIFT_END = time(hour=19, minute=0)


class AttendancePingRequest(BaseModel):
    latitude: float = Field(ge=-90, le=90)
    longitude: float = Field(ge=-180, le=180)


class AttendanceMeOut(BaseModel):
    checked_in: bool
    status: str
    location_status: str
    late_minutes: int
    check_out_at: datetime | None = None
    early_checkout_minutes: int = 0
    warning_text: str | None = None
    check_in_at: datetime | None = None
    last_ping_at: datetime | None = None


class AttendancePolicyOut(BaseModel):
    hotel_id: int
    checkin_start: str
    checkin_end: str
    shift_end: str
    export_mode: str


class AttendancePolicyUpsertRequest(BaseModel):
    checkin_start: str = Field(min_length=4, max_length=5)
    checkin_end: str = Field(min_length=4, max_length=5)
    shift_end: str = Field(min_length=4, max_length=5)
    export_mode: str = Field(default="weekly")
    hotel_id: int | None = None


def _now_utc() -> datetime:
    return datetime.now(timezone.utc)


def _haversine_meters(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    r = 6371000.0
    dlat = radians(lat2 - lat1)
    dlon = radians(lon2 - lon1)
    a = sin(dlat / 2) ** 2 + cos(radians(lat1)) * cos(radians(lat2)) * sin(dlon / 2) ** 2
    c = 2 * asin(sqrt(a))
    return r * c


def _time_to_str(t: time) -> str:
    return t.strftime("%H:%M")


def _parse_hhmm(value: str, field_name: str) -> time:
    txt = (value or "").strip()
    try:
        parsed = datetime.strptime(txt, "%H:%M").time()
    except ValueError:
        raise HTTPException(status_code=400, detail=f"صيغة {field_name} يجب أن تكون HH:MM")
    return parsed


def _policy_export_mode_to_int(mode: str) -> int:
    m = (mode or "weekly").strip().lower()
    if m == "weekly":
        return 0
    if m == "monthly":
        return 1
    raise HTTPException(status_code=400, detail="export_mode يجب أن يكون weekly أو monthly")


def _policy_export_mode_to_str(mode: int) -> str:
    return "monthly" if int(mode or 0) == 1 else "weekly"


def _resolve_hotel_for_policy(db: Session, current_user: User, requested_hotel_id: int | None) -> int:
    role = current_user.role.value if hasattr(current_user.role, "value") else current_user.role
    if role == "admin":
        target = requested_hotel_id
        if target is None:
            if current_user.hotel_id:
                target = current_user.hotel_id
            else:
                first_hotel = db.query(Hotel).order_by(Hotel.id.asc()).first()
                if not first_hotel:
                    raise HTTPException(status_code=404, detail="لا توجد فنادق")
                target = first_hotel.id
    elif role in ["supervisor", "superfv"]:
        if not current_user.hotel_id:
            raise HTTPException(status_code=400, detail="حسابك غير مرتبط بفندق")
        if requested_hotel_id is not None and requested_hotel_id != current_user.hotel_id:
            raise HTTPException(status_code=403, detail="لا تملك صلاحية تعديل سياسة فندق آخر")
        target = current_user.hotel_id
    else:
        raise HTTPException(status_code=403, detail="لا تملك صلاحية إدارة سياسة الحضور")

    exists = db.query(Hotel).filter(Hotel.id == target).first()
    if not exists:
        raise HTTPException(status_code=404, detail="الفندق المطلوب غير موجود")
    return int(target)


def _get_or_create_policy(db: Session, hotel_id: int) -> AttendancePolicy:
    row = db.query(AttendancePolicy).filter(AttendancePolicy.hotel_id == hotel_id).first()
    if row:
        return row

    row = AttendancePolicy(
        hotel_id=hotel_id,
        checkin_start=DEFAULT_CHECKIN_START,
        checkin_end=DEFAULT_CHECKIN_END,
        shift_end=DEFAULT_SHIFT_END,
        export_mode=0,
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


def _policy_out(row: AttendancePolicy) -> AttendancePolicyOut:
    return AttendancePolicyOut(
        hotel_id=row.hotel_id,
        checkin_start=_time_to_str(row.checkin_start),
        checkin_end=_time_to_str(row.checkin_end),
        shift_end=_time_to_str(row.shift_end),
        export_mode=_policy_export_mode_to_str(row.export_mode),
    )


def _early_checkout_minutes(checked_out_at: datetime | None, shift_end_time: time) -> int:
    if not checked_out_at:
        return 0
    end_dt = datetime.combine(checked_out_at.date(), shift_end_time)
    co_naive = checked_out_at.replace(tzinfo=None) if getattr(checked_out_at, "tzinfo", None) else checked_out_at
    minutes = int((end_dt - co_naive).total_seconds() // 60)
    return minutes if minutes > 0 else 0


def _early_checkout_warning_text(checked_out_at: datetime | None, shift_end_time: time) -> tuple[int, str | None]:
    minutes = _early_checkout_minutes(checked_out_at, shift_end_time)
    if minutes >= EARLY_CHECKOUT_WARNING_MINUTES:
        return minutes, f"تحذير: أنهى الدوام مبكرًا قبل النهاية بـ {minutes} دقيقة"
    return minutes, None


@router.get("/attendance/policy", response_model=AttendancePolicyOut)
def attendance_policy_get(
    hotel_id: int | None = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    target_hotel_id = _resolve_hotel_for_policy(db, current_user, hotel_id)
    row = _get_or_create_policy(db, target_hotel_id)
    return _policy_out(row)


@router.put("/attendance/policy", response_model=AttendancePolicyOut)
def attendance_policy_upsert(
    req: AttendancePolicyUpsertRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    target_hotel_id = _resolve_hotel_for_policy(db, current_user, req.hotel_id)
    start_t = _parse_hhmm(req.checkin_start, "checkin_start")
    end_t = _parse_hhmm(req.checkin_end, "checkin_end")
    shift_end_t = _parse_hhmm(req.shift_end, "shift_end")

    if end_t <= start_t:
        raise HTTPException(status_code=400, detail="وقت نهاية نافذة الحضور يجب أن يكون بعد وقت البداية")
    if shift_end_t <= end_t:
        raise HTTPException(status_code=400, detail="وقت نهاية الدوام يجب أن يكون بعد نهاية نافذة الحضور")

    row = _get_or_create_policy(db, target_hotel_id)
    row.checkin_start = start_t
    row.checkin_end = end_t
    row.shift_end = shift_end_t
    row.export_mode = _policy_export_mode_to_int(req.export_mode)
    db.commit()
    db.refresh(row)
    return _policy_out(row)


def _today_session(db: Session, user_id: int) -> AttendanceSession | None:
    return (
        db.query(AttendanceSession)
        .filter(AttendanceSession.user_id == user_id, AttendanceSession.session_date == date.today())
        .order_by(AttendanceSession.id.desc())
        .first()
    )


def _attendance_state(session: AttendanceSession, shift_start: datetime, now_utc: datetime) -> tuple[str, str, int]:
    check_in = session.check_in_at
    check_in_naive = check_in.replace(tzinfo=None) if getattr(check_in, "tzinfo", None) else check_in
    late_minutes = int((check_in_naive - shift_start).total_seconds() // 60)
    if late_minutes < 0:
        late_minutes = 0

    location_status = "in_range"
    if session.last_ping_lat is not None and session.last_ping_lng is not None:
        distance = _haversine_meters(
            float(session.check_in_lat),
            float(session.check_in_lng),
            float(session.last_ping_lat),
            float(session.last_ping_lng),
        )
        if distance > ATTENDANCE_RADIUS_METERS:
            location_status = "out_of_range"

    if session.last_ping_at is None or (now_utc - session.last_ping_at) > timedelta(minutes=ATTENDANCE_STALE_MINUTES):
        location_status = "out_of_range"

    if session.out_of_range_since is not None:
        location_status = "out_of_range"

    if location_status == "out_of_range":
        status = "left_area"
    else:
        status = "present" if late_minutes <= 15 else "late"

    return status, location_status, late_minutes


@router.post("/attendance/check-in", response_model=AttendanceMeOut)
def attendance_check_in(
    req: AttendancePingRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    role = current_user.role.value if hasattr(current_user.role, "value") else current_user.role
    if role in ["admin"]:
        raise HTTPException(status_code=403, detail="هذا الدور لا يسجل حضور")
    if not current_user.hotel_id:
        raise HTTPException(status_code=400, detail="لا يمكن تسجيل الحضور بدون فندق مرتبط")

    policy = _get_or_create_policy(db, current_user.hotel_id)
    now_local = datetime.now()
    now_time = now_local.time().replace(second=0, microsecond=0)
    if now_time < policy.checkin_start:
        raise HTTPException(
            status_code=403,
            detail=f"تسجيل الحضور يبدأ من {_time_to_str(policy.checkin_start)}",
        )

    now_utc = _now_utc()
    session = _today_session(db, current_user.id)

    if session and session.checked_out_at is not None:
        raise HTTPException(status_code=400, detail="تم تسجيل الانصراف اليوم ولا يمكن إعادة تسجيل الحضور")

    if not session:
        session = AttendanceSession(
            user_id=current_user.id,
            hotel_id=current_user.hotel_id,
            session_date=date.today(),
            check_in_at=now_utc,
            check_in_lat=req.latitude,
            check_in_lng=req.longitude,
            last_ping_at=now_utc,
            last_ping_lat=req.latitude,
            last_ping_lng=req.longitude,
            is_active=True,
        )
        db.add(session)
    else:
        session.last_ping_at = now_utc
        session.last_ping_lat = req.latitude
        session.last_ping_lng = req.longitude
        session.is_active = True

    session.out_of_range_since = None
    db.commit()
    db.refresh(session)

    shift_start = datetime.combine(date.today(), policy.checkin_start)
    status, location_status, late_minutes = _attendance_state(session, shift_start, now_utc)
    return AttendanceMeOut(
        checked_in=True,
        status=status,
        location_status=location_status,
        late_minutes=late_minutes,
        check_in_at=session.check_in_at,
        last_ping_at=session.last_ping_at,
    )


@router.post("/attendance/ping", response_model=AttendanceMeOut)
def attendance_ping(
    req: AttendancePingRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    role = current_user.role.value if hasattr(current_user.role, "value") else current_user.role
    if role in ["admin"]:
        raise HTTPException(status_code=403, detail="هذا الدور لا يسجل حضور")

    now_utc = _now_utc()
    session = _today_session(db, current_user.id)
    if not session:
        raise HTTPException(status_code=400, detail="ابدأ الدوام أولاً قبل تحديث الموقع")
    if session.checked_out_at is not None:
        raise HTTPException(status_code=400, detail="تم تسجيل الانصراف اليوم")

    policy = _get_or_create_policy(db, session.hotel_id)

    session.last_ping_at = now_utc
    session.last_ping_lat = req.latitude
    session.last_ping_lng = req.longitude
    session.is_active = True

    distance = _haversine_meters(
        float(session.check_in_lat),
        float(session.check_in_lng),
        float(req.latitude),
        float(req.longitude),
    )
    if distance > ATTENDANCE_RADIUS_METERS:
        session.out_of_range_since = session.out_of_range_since or now_utc
    else:
        session.out_of_range_since = None

    db.commit()
    db.refresh(session)

    shift_start = datetime.combine(date.today(), policy.checkin_start)
    status, location_status, late_minutes = _attendance_state(session, shift_start, now_utc)
    return AttendanceMeOut(
        checked_in=True,
        status=status,
        location_status=location_status,
        late_minutes=late_minutes,
        check_in_at=session.check_in_at,
        last_ping_at=session.last_ping_at,
    )


@router.post("/attendance/check-out", response_model=AttendanceMeOut)
def attendance_check_out(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    role = current_user.role.value if hasattr(current_user.role, "value") else current_user.role
    if role in ["admin"]:
        raise HTTPException(status_code=403, detail="هذا الدور لا يسجل حضور")

    session = _today_session(db, current_user.id)
    if not session:
        raise HTTPException(status_code=400, detail="لا توجد جلسة حضور لليوم")
    if session.checked_out_at is not None:
        raise HTTPException(status_code=400, detail="تم تسجيل الانصراف مسبقًا")

    now_utc = _now_utc()
    session.checked_out_at = now_utc
    session.is_active = False
    db.commit()
    db.refresh(session)

    policy = _get_or_create_policy(db, session.hotel_id)
    shift_start = datetime.combine(date.today(), policy.checkin_start)
    _, _, late_minutes = _attendance_state(session, shift_start, now_utc)
    early_minutes, warning_text = _early_checkout_warning_text(session.checked_out_at, policy.shift_end)
    return AttendanceMeOut(
        checked_in=False,
        status="checked_out",
        location_status="unknown",
        late_minutes=late_minutes,
        check_out_at=session.checked_out_at,
        early_checkout_minutes=early_minutes,
        warning_text=warning_text,
        check_in_at=session.check_in_at,
        last_ping_at=session.last_ping_at,
    )


@router.get("/attendance/me", response_model=AttendanceMeOut)
def attendance_me(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    role = current_user.role.value if hasattr(current_user.role, "value") else current_user.role
    if role in ["admin"]:
        return AttendanceMeOut(
            checked_in=False,
            status="not_applicable",
            location_status="unknown",
            late_minutes=0,
        )

    session = _today_session(db, current_user.id)
    if not session:
        return AttendanceMeOut(
            checked_in=False,
            status="not_started",
            location_status="unknown",
            late_minutes=0,
        )

    if session.checked_out_at is not None:
        policy = _get_or_create_policy(db, session.hotel_id)
        shift_start = datetime.combine(date.today(), policy.checkin_start)
        _, _, late_minutes = _attendance_state(session, shift_start, _now_utc())
        early_minutes, warning_text = _early_checkout_warning_text(session.checked_out_at, policy.shift_end)
        return AttendanceMeOut(
            checked_in=False,
            status="checked_out",
            location_status="unknown",
            late_minutes=late_minutes,
            check_out_at=session.checked_out_at,
            early_checkout_minutes=early_minutes,
            warning_text=warning_text,
            check_in_at=session.check_in_at,
            last_ping_at=session.last_ping_at,
        )

    policy = _get_or_create_policy(db, session.hotel_id)
    now_utc = _now_utc()
    shift_start = datetime.combine(date.today(), policy.checkin_start)
    status, location_status, late_minutes = _attendance_state(session, shift_start, now_utc)
    return AttendanceMeOut(
        checked_in=True,
        status=status,
        location_status=location_status,
        late_minutes=late_minutes,
        check_in_at=session.check_in_at,
        last_ping_at=session.last_ping_at,
    )


def _repair_minutes(report: MaintenanceReport) -> float | None:
    if not report.reported_at or not report.completed_at:
        return None

    start = report.reported_at
    end = report.completed_at

    # Normalize potential timezone mismatch between DB drivers.
    if start.tzinfo is None and end.tzinfo is not None:
        start = start.replace(tzinfo=end.tzinfo)
    if end.tzinfo is None and start.tzinfo is not None:
        end = end.replace(tzinfo=start.tzinfo)

    delta = end - start
    return delta.total_seconds() / 60


@router.get("/overview", response_model=DashboardOverviewOut)
def dashboard_overview(
    hotel_id: int | None = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    role = current_user.role.value if hasattr(current_user.role, "value") else current_user.role

    report_query = db.query(MaintenanceReport)
    room_query = db.query(Room)
    hotel_query = db.query(Hotel)

    if role != "admin":
        report_query = report_query.filter(MaintenanceReport.hotel_id == current_user.hotel_id)
        room_query = room_query.filter(Room.hotel_id == current_user.hotel_id)
        hotel_query = hotel_query.filter(Hotel.id == current_user.hotel_id)
    elif hotel_id is not None:
        exists = db.query(Hotel).filter(Hotel.id == hotel_id).first()
        if not exists:
            raise HTTPException(status_code=404, detail="الفندق المطلوب غير موجود")
        report_query = report_query.filter(MaintenanceReport.hotel_id == hotel_id)
        room_query = room_query.filter(Room.hotel_id == hotel_id)
        hotel_query = hotel_query.filter(Hotel.id == hotel_id)

    reports = report_query.all()
    rooms = room_query.all()
    hotels = hotel_query.all()

    total_faults = len(reports)
    open_faults = len([r for r in reports if r.status not in [MaintenanceStatus.completed, MaintenanceStatus.verified]])
    waiting_parts_faults = len([r for r in reports if r.status == MaintenanceStatus.waiting_parts])
    completed_faults = len([r for r in reports if r.status == MaintenanceStatus.completed])
    verified_faults = len([r for r in reports if r.status == MaintenanceStatus.verified])

    # Fastest technician: minimum average repair time among technicians with at least one completed report.
    tech_durations: Dict[int, List[float]] = defaultdict(list)
    for report in reports:
        if not report.assigned_to_id:
            continue
        minutes = _repair_minutes(report)
        if minutes is None:
            continue
        tech_durations[report.assigned_to_id].append(minutes)

    fastest_technician = None
    if tech_durations:
        tech_avgs = []
        for user_id, values in tech_durations.items():
            avg = sum(values) / len(values)
            tech_avgs.append((user_id, avg, len(values)))

        fastest_user_id, fastest_avg, resolved = sorted(tech_avgs, key=lambda x: x[1])[0]
        user = db.query(User).filter(User.id == fastest_user_id).first()
        if user:
            fastest_technician = TechnicianPerformanceOut(
                user_id=user.id,
                full_name=user.full_name,
                avg_repair_minutes=round(fastest_avg, 2),
                resolved_reports=resolved,
            )

    # Slowest hotel: maximum average repair time among hotels with completed reports.
    hotel_durations: Dict[int, List[float]] = defaultdict(list)
    for report in reports:
        minutes = _repair_minutes(report)
        if minutes is None:
            continue
        hotel_durations[report.hotel_id].append(minutes)

    slowest_hotel = None
    if hotel_durations:
        hotel_avgs = []
        for hotel_id, values in hotel_durations.items():
            avg = sum(values) / len(values)
            hotel_avgs.append((hotel_id, avg, len(values)))

        slowest_hotel_id, slowest_avg, resolved = sorted(hotel_avgs, key=lambda x: x[1], reverse=True)[0]
        hotel = db.query(Hotel).filter(Hotel.id == slowest_hotel_id).first()
        if hotel:
            slowest_hotel = HotelPerformanceOut(
                hotel_id=hotel.id,
                hotel_name=hotel.name,
                avg_repair_minutes=round(slowest_avg, 2),
                resolved_reports=resolved,
            )

    # Uptime = ready rooms / total rooms.
    total_rooms = len(rooms)
    ready_rooms = len([r for r in rooms if r.status == RoomStatus.ready])
    uptime_percent = round((ready_rooms / total_rooms) * 100, 2) if total_rooms else 0.0

    overall_uptime = RoomUptimeOut(
        total_rooms=total_rooms,
        ready_rooms=ready_rooms,
        uptime_percent=uptime_percent,
    )

    hotels_uptime: List[RoomUptimeOut] = []
    for hotel in hotels:
        hotel_rooms = [r for r in rooms if r.hotel_id == hotel.id]
        h_total = len(hotel_rooms)
        h_ready = len([r for r in hotel_rooms if r.status == RoomStatus.ready])
        h_uptime = round((h_ready / h_total) * 100, 2) if h_total else 0.0
        hotels_uptime.append(
            RoomUptimeOut(
                hotel_id=hotel.id,
                hotel_name=hotel.name,
                total_rooms=h_total,
                ready_rooms=h_ready,
                uptime_percent=h_uptime,
            )
        )

    return DashboardOverviewOut(
        faults=DashboardCountOut(
            total=total_faults,
            open=open_faults,
            waiting_parts=waiting_parts_faults,
            completed=completed_faults,
            verified=verified_faults,
        ),
        fastest_technician=fastest_technician,
        slowest_hotel=slowest_hotel,
        rooms_uptime=overall_uptime,
        hotels_uptime=hotels_uptime,
    )


@router.get("/attendance", response_model=List[AttendanceRowOut])
def dashboard_attendance(
    hotel_id: int | None = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    role = current_user.role.value if hasattr(current_user.role, "value") else current_user.role

    target_hotel_id = hotel_id
    if role != "admin":
        target_hotel_id = current_user.hotel_id
    elif target_hotel_id is not None:
        exists = db.query(Hotel).filter(Hotel.id == target_hotel_id).first()
        if not exists:
            raise HTTPException(status_code=404, detail="الفندق المطلوب غير موجود")

    users_q = db.query(User).filter(User.is_active.is_(True))
    if target_hotel_id is not None:
        users_q = users_q.filter(User.hotel_id == target_hotel_id)

    if role in ["supervisor", "superfv"]:
        users_q = users_q.filter(
            User.role.notin_([
                UserRole.admin,
                UserRole.accountant,
                UserRole.supervisor,
                UserRole.superfv,
            ])
        )

    users = users_q.all()
    if not users:
        return []

    user_ids = [u.id for u in users]
    today = date.today()

    sessions = (
        db.query(AttendanceSession)
        .filter(AttendanceSession.user_id.in_(user_ids), AttendanceSession.session_date == today)
        .all()
    )
    sessions_by_user = {s.user_id: s for s in sessions}

    hotel_ids = sorted({s.hotel_id for s in sessions if s.hotel_id is not None})
    policies_map: Dict[int, AttendancePolicy] = {}
    if hotel_ids:
        policy_rows = db.query(AttendancePolicy).filter(AttendancePolicy.hotel_id.in_(hotel_ids)).all()
        policies_map = {p.hotel_id: p for p in policy_rows}

    now_local = datetime.now()
    now_utc = _now_utc()

    out: List[AttendanceRowOut] = []
    for u in users:
        session = sessions_by_user.get(u.id)
        check_in = session.check_in_at if session else None

        shift_start_time = DEFAULT_CHECKIN_START
        shift_end_time = DEFAULT_SHIFT_END
        if session and session.hotel_id in policies_map:
            policy = policies_map[session.hotel_id]
            shift_start_time = policy.checkin_start
            shift_end_time = policy.shift_end
        shift_start = datetime.combine(today, shift_start_time)
        before_shift_start = now_local < shift_start

        if session is None:
            status_label = "not_started" if before_shift_start else "absent"
            out.append(
                AttendanceRowOut(
                    user_id=u.id,
                    full_name=u.full_name,
                    role=u.role.value if hasattr(u.role, "value") else str(u.role),
                    check_in_at=None,
                    check_out_at=None,
                    location_status="unknown",
                    status=status_label,
                    late_minutes=0,
                    early_checkout_minutes=0,
                    warning_text=None,
                )
            )
            continue

        if session.checked_out_at is not None:
            check_in_naive = check_in.replace(tzinfo=None) if getattr(check_in, 'tzinfo', None) else check_in
            late_minutes = int((check_in_naive - shift_start).total_seconds() // 60)
            if late_minutes < 0:
                late_minutes = 0
            status_label = "checked_out"
            location = "unknown"
            early_minutes, warning_text = _early_checkout_warning_text(session.checked_out_at, shift_end_time)
        else:
            status_label, location, late_minutes = _attendance_state(session, shift_start, now_utc)
            early_minutes, warning_text = 0, None

        out.append(
            AttendanceRowOut(
                user_id=u.id,
                full_name=u.full_name,
                role=u.role.value if hasattr(u.role, "value") else str(u.role),
                check_in_at=check_in,
                check_out_at=session.checked_out_at,
                location_status=location,
                status=status_label,
                late_minutes=late_minutes,
                early_checkout_minutes=early_minutes,
                warning_text=warning_text,
            )
        )

    status_order = {"present": 0, "late": 1, "left_area": 2, "checked_out": 3, "not_started": 4, "absent": 5}
    out.sort(key=lambda x: (status_order.get(x.status, 99), x.full_name))
    return out


@router.get("/attendance/export")
def attendance_export_xlsx(
    period: str = "weekly",
    hotel_id: int | None = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    role = current_user.role.value if hasattr(current_user.role, "value") else current_user.role
    if role not in ["admin", "supervisor", "superfv", "accountant"]:
        raise HTTPException(status_code=403, detail="لا تملك صلاحية تصدير الحضور")

    p = (period or "weekly").strip().lower()
    if p not in ["weekly", "monthly"]:
        raise HTTPException(status_code=400, detail="period يجب أن يكون weekly أو monthly")

    today = date.today()
    if p == "weekly":
        from_date = today - timedelta(days=today.weekday())
        to_date = from_date + timedelta(days=6)
    else:
        from_date = today.replace(day=1)
        if today.month == 12:
            next_month = today.replace(year=today.year + 1, month=1, day=1)
        else:
            next_month = today.replace(month=today.month + 1, day=1)
        to_date = next_month - timedelta(days=1)

    q = db.query(AttendanceSession).filter(
        AttendanceSession.session_date >= from_date,
        AttendanceSession.session_date <= to_date,
    )

    target_hotel_id = hotel_id
    if role != "admin":
        target_hotel_id = current_user.hotel_id
    elif target_hotel_id is not None:
        exists = db.query(Hotel).filter(Hotel.id == target_hotel_id).first()
        if not exists:
            raise HTTPException(status_code=404, detail="الفندق المطلوب غير موجود")

    if target_hotel_id is not None:
        q = q.filter(AttendanceSession.hotel_id == target_hotel_id)

    sessions = q.order_by(AttendanceSession.session_date.asc(), AttendanceSession.check_in_at.asc()).all()

    user_ids = sorted({s.user_id for s in sessions})
    users_map = {}
    if user_ids:
        users_map = {u.id: u for u in db.query(User).filter(User.id.in_(user_ids)).all()}

    hotel_ids = sorted({s.hotel_id for s in sessions})
    policies_map = {}
    if hotel_ids:
        policies_map = {p.hotel_id: p for p in db.query(AttendancePolicy).filter(AttendancePolicy.hotel_id.in_(hotel_ids)).all()}
    hotels_map = {}
    if hotel_ids:
        hotels_map = {h.id: h for h in db.query(Hotel).filter(Hotel.id.in_(hotel_ids)).all()}

    wb = Workbook()
    ws = wb.active
    ws.title = "attendance"
    ws.sheet_view.rightToLeft = True
    rtl_align = Alignment(horizontal="right")
    headers = [
        "التاريخ",
        "الفندق",
        "الموظف",
        "الدور",
        "دخول",
        "خروج",
        "الحالة",
        "الموقع",
        "دقائق التأخير",
        "تحذير",
    ]
    # Some viewers (notably web spreadsheets) may ignore sheet RTL metadata,
    # so we physically reverse columns to preserve RTL reading order.
    ws.append(list(reversed(headers)))
    for cell in ws[1]:
        cell.alignment = rtl_align

    for s in sessions:
        user = users_map.get(s.user_id)
        hotel = hotels_map.get(s.hotel_id)
        policy = policies_map.get(s.hotel_id)
        checkin_start = policy.checkin_start if policy else DEFAULT_CHECKIN_START
        shift_end = policy.shift_end if policy else DEFAULT_SHIFT_END
        shift_start = datetime.combine(s.session_date, checkin_start)

        status, location, late_minutes = _attendance_state(s, shift_start, _now_utc()) if s.checked_out_at is None else ("checked_out", "unknown", max(0, int(((s.check_in_at.replace(tzinfo=None) if getattr(s.check_in_at, "tzinfo", None) else s.check_in_at) - shift_start).total_seconds() // 60)))
        early_minutes, warning_text = _early_checkout_warning_text(s.checked_out_at, shift_end)

        row_values = [
            str(s.session_date),
            hotel.name if hotel else f"فندق {s.hotel_id}",
            user.full_name if user else f"#{s.user_id}",
            (user.role.value if hasattr(user.role, "value") else str(user.role)) if user else "-",
            s.check_in_at.strftime("%Y-%m-%d %H:%M:%S") if s.check_in_at else "-",
            s.checked_out_at.strftime("%Y-%m-%d %H:%M:%S") if s.checked_out_at else "-",
            status,
            location,
            late_minutes,
            warning_text or "",
        ]
        ws.append(list(reversed(row_values)))

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=10):
        for cell in row:
            cell.alignment = rtl_align

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    fname = f"attendance_{p}_{from_date.isoformat()}_{to_date.isoformat()}.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{fname}"'}
    return StreamingResponse(out, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)
